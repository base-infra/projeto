from io import BytesIO
from django.http import HttpResponse, Http404
from django.shortcuts import get_object_or_404
from django.forms.models import model_to_dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..models import Interessado, DadosPdfValidado
from datetime import datetime

def _status_from(value, partial=False):
    if value:
        return "Aprovado"
    if value is None or partial:
        return "Revisar"
    return "Em Revisão"

def exportar_checklist(request, interessado_id, pdf_id):
    print(f"\n[DEBUG] --- exportar_checklist ---")
    print(f"[DEBUG] Interessado ID: {interessado_id}")
    print(f"[DEBUG] PDF ID: {pdf_id}")

    interessado = get_object_or_404(Interessado, pk=interessado_id)
    checklist = get_object_or_404(DadosPdfValidado, pk=pdf_id, interessado=interessado)

    print(f"[DEBUG] Interessado encontrado: {interessado}")
    print(f"[DEBUG] PDF encontrado: {checklist}")

    c = model_to_dict(checklist)
    checklist_data = {
        **c,
        "interessado": {
            "id": interessado.id,
            "nro_processo": interessado.nro_processo,
            "nome": interessado.nome,
        },
        "nome_br": getattr(checklist, "nome_br", ""),
        "sentido": getattr(checklist, "sentido", ""),
        "classificacao": getattr(checklist, "classificacao", ""),
    }

    print(f"[DEBUG] checklist_data.keys(): {list(checklist_data.keys())}")
    print(f"[DEBUG] Localizacao: {checklist_data.get('localizacao')}")
    print(f"[DEBUG] Kilometragem inicio: {checklist_data.get('kilometragem_inicio')}")
    print(f"[DEBUG] Kilometragem fim: {checklist_data.get('kilometragem_fim')}")
    print(f"[DEBUG] Nome BR: {checklist_data.get('nome_br')}")
    print(f"[DEBUG] Classificacao: {checklist_data.get('classificacao')}")

    # Criar Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist"

    hfill = PatternFill("solid", fgColor="00264D")
    hfont = Font(color="FFFFFF", bold=True)
    b = Border(left=Side(style="thin"), right=Side(style="thin"),
               top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:H1")
    ws["A1"] = "BASEINFRA PROJETOS E CONSULTORIA"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    ws.merge_cells("A2:H2")
    ws["A2"] = f"CHECKLIST DE DOCUMENTOS – {str(checklist_data['classificacao']).upper()}"
    ws["A2"].alignment = center

    ws.append([])
    ws.append(["Protocolo","BIP","Revisão","Interessado"])
            #    "Localização de Acesso","Classificação","BR/Sentido/Tipo","Nº Folha"
    for ccell in ws[4]:
        ccell.fill, ccell.font, ccell.border, ccell.alignment = hfill, hfont, b, center

    ws.append([
        checklist_data["interessado"]["nro_processo"],
        f"BIP-001-{checklist_data['interessado']['id']}",
        "01",
        checklist_data["interessado"]["nome"].upper(),
        # str(checklist_data.get("localizacao","")).upper(),
        # str(checklist_data.get("classificacao","")).upper(),
        # f"{str(checklist_data.get('nome_br','')).upper()} / {str(checklist_data.get('sentido','')).upper()} / {str(checklist_data.get('classificacao','')).upper()}",
        # str(checklist_data.get("nro_prancha","")).upper(),
    ])
    for ccell in ws[5]:
        ccell.border = b

    ws.append([])
    ws.append(["Descrição do Processo", str(checklist_data.get("descricao_processo","")).upper()])
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=8)
    for ccell in ws[7]:
        ccell.border = b

    ws.append([])
    start_header_row = ws.max_row + 1
    ws.append(["ITEM","TIPO DO DOCUMENTO/PROJETO NECESSÁRIOS","REVISÃO","SITUAÇÃO","RESULTADO"])
    for ccell in ws[start_header_row]:
        ccell.fill, ccell.font, ccell.border, ccell.alignment = hfill, hfont, b, center

    itens = [
        (1, "Localização do acesso conforme SRE vigente", "Rev. 01", _status_from(checklist_data.get("localizacao")), str(checklist_data.get("localizacao", "")).upper()),
        (2, "KM+M do início, final e do eixo do acesso", "Rev. 01",
        _status_from(checklist_data.get("kilometragem_inicio") and checklist_data.get("kilometragem_fim"),
                    partial=(checklist_data.get("kilometragem_inicio") and checklist_data.get("kilometragem_fim") is None)),
        f"{str(checklist_data.get('kilometragem_inicio','')).upper()} / {str(checklist_data.get('kilometragem_fim','')).upper()}"),
        (3, "Código e trecho da rodovia", "Rev. 01", _status_from(checklist_data.get("nome_br")), str(checklist_data.get("nome_br", "")).upper()),
        (4, "Coordenadas georreferenciadas em UTM SIRGAS 2000", "Rev. 01",
        _status_from(checklist_data.get("coordenadas_georreferenciais_e") and checklist_data.get("coordenadas_georreferenciais_n")),
        f"{str(checklist_data.get('coordenadas_georreferenciais_e', '')).upper()} / {str(checklist_data.get('coordenadas_georreferenciais_n', '')).upper()}"),
        (5, "O traçado da faixa de domínio", "Rev. 01", _status_from(checklist_data.get("tracado_dominio")), str(checklist_data.get("tracado_dominio", "")).upper()),
        (6, "Cotas e textos legíveis", "Rev. 01", _status_from(checklist_data.get("cotas_legiveis")), str(checklist_data.get("cotas_legiveis", "")).upper()),
        (7, "Verificação da escala", "Rev. 01", _status_from(checklist_data.get("escala_pdf")), str(checklist_data.get("escala_pdf", "")).upper()),
        (8, "Conferir se o memorial está de acordo com o projeto", "Rev. 01", _status_from(checklist_data.get("memorial_pdf")), str(checklist_data.get("memorial_pdf", "")).upper()),
        (9, "Conferir larguras das pistas pelo DNIT", "Rev. 01", _status_from(checklist_data.get("largura_pista_dnit")), str(checklist_data.get("largura_pista_dnit", "")).upper()),
        (10, "Verificar se as legendas estão de acordo com o projeto", "Rev. 01", _status_from(checklist_data.get("legenda_correta")), str(checklist_data.get("legenda_correta", "")).upper()),
        (11, "Verificar anotações e notas do projeto", "Rev. 01", _status_from(checklist_data.get("anotacao_nota")), str(checklist_data.get("anotacao_nota", "")).upper()),
        (12, "Verificar siglas e abreviações", "Rev. 01", _status_from(checklist_data.get("sigla_abreviacao")), str(checklist_data.get("sigla_abreviacao", "")).upper()),
        (13, "Localização do acesso, quilometragem e prefixo", "Rev. 01",
        _status_from(checklist_data.get("localizacao") and checklist_data.get("kilometragem_inicio") and checklist_data.get("kilometragem_fim")),
        f"{str(checklist_data.get('localizacao', '')).upper()} / {str(checklist_data.get('kilometragem_inicio', '')).upper()}/ {str(checklist_data.get('kilometragem_fim','')).upper()}"),
        (14, "Verificar se o carimbo apresenta as informações corretas", "Rev. 01", _status_from(checklist_data.get("carimbo_correto")), str(checklist_data.get("carimbo_correto", "")).upper()),
        (15, "Limites de propriedade em relação à rodovia", "Rev. 01", _status_from(checklist_data.get("limite_propriedade")), str(checklist_data.get("limite_propriedade", "")).upper()),
        (16, "Delimitação da faixa de domínio e faixa não edificante", "Rev. 01", _status_from(checklist_data.get("delimitacao_dominio_nao_edificante")), str(checklist_data.get("delimitacao_dominio_nao_edificante", "")).upper()),
        (17, "ART - Anotação de Responsabilidade Técnica", "Rev. 01", _status_from(checklist_data.get("art_pdf")), str(checklist_data.get("art_pdf", "")).upper()),
    ]


    for item, desc, rev, status, resultado in itens:
        ws.append([item, desc, rev, status, resultado])
        for ccell in ws[ws.max_row]:
            ccell.border = b
            if ccell.column == 1:
                ccell.alignment = center

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    for col in "EFGH":
        ws.column_dimensions[col].width = 18

    print(f"[DEBUG] Total linhas Excel: {ws.max_row}")
    print(f"[DEBUG] Total colunas Excel: {ws.max_column}")


    # Nome do cliente formatado (substitui espaços por _ e põe em maiúsculas)
    nome_cliente = interessado.nome.strip().upper().replace(" ", "_")

    # Obtém data e hora atual no formato desejado
    data_hora_str = datetime.now().strftime("%Y%m%d_%H%M")

    # Formata o nome do ficheiro com data e hora
    filename = f"{nome_cliente}_checklist_{interessado_id}_{pdf_id}_{data_hora_str}.xlsx" 

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    file_size = len(buf.getvalue())
    print(f"[DEBUG] Tamanho do ficheiro gerado: {file_size} bytes")

    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
